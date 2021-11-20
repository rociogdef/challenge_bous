from openpyxl import load_workbook
import csv
import sys




try:
   wb = load_workbook(filename = sys.argv[1])
except Exception as ex:
   print("No es posible leer el archivo de entrada")
   exit(1)
    

ws=wb.active


colname=list(ws.rows)[2]  #nombre de las columnas primer renglon 
tiposcol=list(ws.rows)[3]


def valida_dato(dato,valor):
   if dato.lower().strip() == valor:
     return True
   return False

l_colname=['examen técnico desarrollador backend','cliente','# contrato','fecha de compra','ciudad','empresa','listado de adeudos']
l_tipos=['(alfabético)','(alfanumérico)','(fecha)','(alfabético)','(alfanumérico)','(numérico)']
def Valida_encabezado(ws):
   if not valida_dato(ws['B1'].value,l_colname[0]):
      print("Existe un error en el primer titulo")
      return False
   if not valida_dato(ws['B2'].value,l_colname[6]):
      print("Existe un error en el segudno titulo")
      return False
   for i in range(1,5):
      if not valida_dato(colname[i].value,l_colname[i]): 
         print("Existe un error en el nombre de la columna "+colname[i].value) 
         return False
   for i in range(1,5):
      if not valida_dato(tiposcol[i].value,l_tipos[i-1]): 
         print("Existe un error en el tipo correspondiente a la columna "+tiposcol[i].value) 
         return False         
   return True

#Genera el archivo .csv sin el encabezado
def Transforma_csv(sh): 
   with open('adeudos_aux.csv', 'w', newline="", encoding='utf-8') as f:  
      c = csv.writer(f)    
      for r in sh.iter_rows(min_row=5, min_col=2, max_col=7):
            c.writerow([cell.value for cell in r])
       
if __name__ == '__main__':
  if not Valida_encabezado(ws):
     exit(1)
  Transforma_csv(ws)   
 